import os
import zipfile
import time
import base64
import tempfile
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import eel

# Inicializa a pasta web do Eel
eel.init('web')

@eel.expose
def selecionar_pasta_destino():
    """Abre uma janela nativa do Windows para o usuário escolher onde salvar."""
    root = tk.Tk()
    root.attributes("-topmost", True) # Mantém a janela no topo
    root.withdraw() # Oculta a janela principal do Tkinter
    folder_path = filedialog.askdirectory(title="Selecione onde salvar os arquivos gerados")
    root.destroy()
    return folder_path

@eel.expose
def abrir_caminho(caminho):
    """Abre o arquivo ou pasta no Windows Explorer."""
    try:
        os.startfile(caminho)
    except Exception as e:
        print(f"Erro ao abrir arquivo: {e}")

@eel.expose
def processar_arquivo_base64(b64_data, output_dir):
    """Recebe o arquivo do navegador, salva temporariamente e inicia o processamento."""
    try:
        # Decodifica o arquivo recebido via Drag & Drop do navegador
        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        with os.fdopen(temp_fd, 'wb') as f:
            f.write(base64.b64decode(b64_data))
        
        # Inicia a transição de dados
        processar_planilhas_core(temp_path, output_dir)
        
        # Limpeza do arquivo temporário
        os.remove(temp_path)
    except Exception as e:
        eel.erro_processamento(f"Erro na conversão do arquivo: {str(e)}")()

def processar_planilhas_core(input_file, output_dir):
    """Lógica central de extração, transposição e formatação do Excel."""
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
        sheet_names = wb.sheetnames
        total_sheets = len(sheet_names)
        
        arquivos_gerados = []
        zip_path = os.path.join(output_dir, "todas_as_planilhas.zip")
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for idx, name in enumerate(sheet_names):
                # Atualiza a barra de progresso no frontend via Eel
                porcentagem = int(((idx + 1) / total_sheets) * 100)
                eel.atualizar_progresso(porcentagem, f"Processando: {name}")()
                
                ws_in = wb[name]
                if ws_in.max_row < 1:
                    continue
                    
                headers = [str(cell.value).strip().upper() if cell.value else "" for cell in ws_in[1]]
                
                def obter_index(nomes_possiveis):
                    for n in nomes_possiveis:
                        if n in headers:
                            return headers.index(n) + 1
                    return None
                    
                idx_desc = obter_index(["DESCRIÇÃO", "DESCRICAO"])
                idx_pos = obter_index(["POSIÇÃO", "POSICAO"])
                idx_tipo = obter_index(["TIPO"])
                idx_tam = obter_index(["TAMANHO"])
                
                # Ignora a planilha se não possuir as colunas obrigatórias
                if not any([idx_desc, idx_pos, idx_tipo, idx_tam]):
                    continue
                    
                dados_colunas = []
                for r in range(2, ws_in.max_row + 1):
                    desc_v = ws_in.cell(row=r, column=idx_desc).value if idx_desc else ""
                    pos_v = ws_in.cell(row=r, column=idx_pos).value if idx_pos else ""
                    tipo_v = ws_in.cell(row=r, column=idx_tipo).value if idx_tipo else ""
                    tam_v = ws_in.cell(row=r, column=idx_tam).value if idx_tam else ""
                    
                    if all(v is None or v == "" for v in [desc_v, pos_v, tipo_v, tam_v]):
                        continue
                        
                    # Conversão estrutural de tipos
                    try:
                        pos_v = int(float(pos_v)) if pos_v is not None and pos_v != "" else ""
                    except ValueError:
                        pos_v = str(pos_v)
                        
                    try:
                        tam_v = int(float(tam_v)) if tam_v is not None and tam_v != "" else ""
                    except ValueError:
                        tam_v = str(tam_v)
                        
                    dados_colunas.append({
                        "desc": str(desc_v) if desc_v is not None else "",
                        "pos": pos_v,
                        "tipo": str(tipo_v) if tipo_v is not None else "",
                        "tam": tam_v
                    })
                    
                if not dados_colunas:
                    continue
                    
                wb_out = openpyxl.Workbook()
                ws_out = wb_out.active
                ws_out.title = name[:30]
                
                # Preenchimento estruturado
                for c_idx, col in enumerate(dados_colunas, start=1):
                    # Linha 1
                    cell_desc = ws_out.cell(row=1, column=c_idx, value=col["desc"])
                    cell_desc.data_type = 's'
                    # Linha 2
                    cell_pos = ws_out.cell(row=2, column=c_idx, value=col["pos"])
                    if isinstance(col["pos"], int): cell_pos.data_type = 'n'
                    # Linha 3
                    cell_tipo = ws_out.cell(row=3, column=c_idx, value=col["tipo"])
                    cell_tipo.data_type = 's'
                    # Linha 4
                    cell_tam = ws_out.cell(row=4, column=c_idx, value=col["tam"])
                    if isinstance(col["tam"], int): cell_tam.data_type = 'n'
                
                # Configurações de UI do Excel
                ws_out.freeze_panes = 'A2'
                max_letter = get_column_letter(len(dados_colunas))
                ws_out.auto_filter.ref = f"A1:{max_letter}1"
                
                # Formatação Condicional (Linha 5 adiante)
                fill_color = "97A4F1"
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                for c_idx in range(1, len(dados_colunas) + 1):
                    letra = get_column_letter(c_idx)
                    regra_formula = f"LEN({letra}5)>{letra}$4"
                    regra = FormulaRule(formula=[regra_formula], stopIfTrue=True, fill=fill)
                    ws_out.conditional_formatting.add(f"{letra}5:{letra}1000", regra)
                    
                nome_limpo = "".join([c for c in name if c.isalnum() or c in " _-"]).strip()
                file_name = f"{nome_limpo}.xlsx"
                full_path = os.path.join(output_dir, file_name)
                
                wb_out.save(full_path)
                arquivos_gerados.append({"nome": file_name, "caminho": full_path})
                zipf.write(full_path, file_name)
                
                time.sleep(0.1) # Pausa mínima para fluidez visual da barra de carregamento
                
        # Envia comando de finalização para o frontend
        eel.finalizar_processamento(arquivos_gerados, zip_path, output_dir)()
        
    except Exception as e:
        eel.erro_processamento(str(e))()

if __name__ == "__main__":
    # Inicia a aplicação utilizando a janela padrão do Chrome/Edge do sistema
    eel.start('index.html', size=(600, 750), port=0)